import sys                      # 更改编码UTF8
# import time                   # 强制等待加载
import requests                 # 请求网页
import json                     # 保存为JSON
# import pandas as pd           # 如果用DataFrame
from bs4 import BeautifulSoup   # 解析网页
# import lxml                   # LXML Parser
import xlsxwriter               # 如果用XlsxWriter
import openpyxl
# from openpyxl import Workbook   
from openpyxl.utils import get_column_letter # 设置Excel行款列宽用

import smtplib
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

PATH = 'src/demo_16'
URL_BASE = 'http://www.ajztb.com'
URL_SDIR = 'jyxx'
URL_FILE = 'moreinfo.html'

class Date:
    year = 2000
    month = 1
    day = 1
    def __init__(self, year_, month_, day_):
        self.year  = year_ 
        self.month = month_
        self.day   = day_

    def toString(self, seperator='-'):
        y_str = str(self.year)
        m_str = str(self.month)
        d_str = str(self.day)
        if(len(m_str)==1): m_str = '0'+m_str
        if(len(d_str)==1): d_str = '0'+d_str
        r_str = y_str+seperator+m_str+seperator+d_str 
        return r_str
    
    def compareTo(self,date):
        # Return:
        #       True if self > date 
        #       False otherise 
        #       None if equal
        if  (self.year > date.year): return True
        elif(self.year < date.year): return False
        else:
            if  (self.month > date.month): return True
            elif(self.month < date.month): return False
            else:
                if  (self.day > date.day): return True
                elif(self.day < date.day): return False
                else: return None

    def iter_date_to(self, date):
        # Swap if ordering is not satisfied
        if(self.compareTo(date) == False):
            temp = self 
            self = date 
            date = temp

        # since date is "inclusive"
        date.day -= 1 

        # process for date list
        rtn = []
        big_month   = [1,3,5,7,8,10,12] 
        small_month = [4,6,9,11]
        while not (self.compareTo(date) is None):
            temp = Date(self.year, self.month, self.day)
            rtn = rtn + [temp] 
            self.day -= 1
            if(self.day == 0):
                self.month -= 1
                if(self.month in big_month):     self.day=31
                elif(self.month in small_month): self.day=30
                else: 
                    if(self.year % 4 == 0): self.day = 29
                    else:                   self.day = 28
                if(self.month == 0):
                    self = Date(self.year-1, 12, 31)
        return rtn

    def prev(self):
        self.day -= 1
        if(self.day == 0):
            big_month   = [1,3,5,7,8,10,12] 
            small_month = [4,6,9,11]
            self.month -= 1
            if(self.month in big_month):     self.day=31
            elif(self.month in small_month): self.day=30
            else: 
                if(self.year % 4 == 0): self.day = 29
                else:                   self.day = 28
            if(self.month == 0):
                self = Date(self.year-1, 12, 31)
        return self

# ========
# Mail utilities
PATH = 'src/demo_16'
HOST = {
    'server'    : 'smtp.163.com',
    'account'   : 'suowei_h_temp@163.com', #myTempMai willBeDeactivated
    'password'  : 'JUIFZIYOWSJFQYIG'       #husuowei200029
}
MAIL = {
    'sender'    : 'suowei_h_temp@163.com',
    'receivers' : ['suowei.h@icloud.com'],
    'title'     : 'TEST - 爬到新信息',
    'message'   : '爬到如下新信息:\n',
    'data'      : '\t xxxx, yyyyy, zzzz \n' * 10 
}

# =========
# Send mail
def sendEmail163(data=MAIL['data'],atts=['result.xlsx']):
    mail_host = HOST['server']                     # SMTP服务器(这里用的163)
    mail_user = HOST['account']                    # 用户名（这里用的临时邮箱，很可能会注销吊）
    mail_pass = HOST['password']                   # 授权码（密码hu****029）
    sender    = MAIL['sender']                     # 发件人邮箱
    receivers = MAIL['receivers']                  # 接收邮件列表
    title     = MAIL['title']                      # 邮件主题
    content   = MAIL['message'] + data             # 邮件内容

    message = MIMEMultipart()
    msg_con = MIMEText(content) 

    message['From'] = "{}".format(sender)
    message['To'] = ",".join(receivers)
    message['Subject'] = title
    message.attach(msg_con)
    message["Accept-Language"]="zh-CN" # 设置消息为中文
    message["Accept-Charset"]="utf-8"  # 指定编码

    for att_name in atts:              # 添加附件
        file_type = 'excel'
        att1 = MIMEText(open(att_name, 'rb').read(), 'base64', 'utf-8')
        att1["Content-Type"] = 'application/octet-stream'
        att1["Content-Disposition"] = 'attachment; filename=' + att_name
        message.attach(att1)               
 
    try:
        smtpObj = smtplib.SMTP_SSL(mail_host, 465)  # 启用SSL发信, 端口一般是465
        smtpObj.login(mail_user, mail_pass)         # 登录验证 163 邮箱
        smtpObj.sendmail(sender, receivers, message.as_string())# 发送
        print("mail has been send successfully.")

    except smtplib.SMTPException as e:
        print(e)

# ===========
# File I/O Utilities

def save_toHtml(text,path=PATH+'/result.html'):
    with open(path, 'w+') as f:
        f.write(text)
        f.close()

def save_toJson(text, path=PATH+'/result.json'):
    s = json.dumps(text, indent=4, ensure_ascii=False)
    with open(path, "w+") as f:
        f.write(s)
        f.close()

def save_toExcel(text,path=PATH+'/result.xlsx',wk_sheet="main"):
    
    # open workbook
    workbook = openpyxl.load_workbook(path)

    if(wk_sheet in workbook.sheetnames): 
        ws = workbook[wk_sheet]
        workbook.remove(ws)
    
    if('Placeholder' in workbook.sheetnames): 
        ws = workbook['Placeholder']
        workbook.remove(ws)
    
    worksheet = workbook.create_sheet(wk_sheet)
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 120
    worksheet.column_dimensions['C'].width = 120
    # worksheet.column_dimensions[ 0 ].height= 30

    # Attempt 1
    # for col, val in enumerate(text, start=1):
    #     worksheet.cell(row=2, column=col).value = val

    # Attempt 2 
    # for row_num, data in enumerate(text):
    #     worksheet.write_row(row_num, 0, data)

    # Attempt 3
    # print(list(text[1]))
    for i in range(0,len(text)):
        for j in range(0,len(text[i])):
            worksheet.cell(row=i+1,column=j+1).value = list(text[i])[j]
    
    # close work book
    workbook.save(path)

def clear_Excel(path, placeholder="Placeholder"):

    workbook  = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet(placeholder)
    workbook.close()

    # workbook = openpyxl.load_workbook(path)
    # print("=" * 100)
    # workbook.create_sheet(placeholder)
    

    # for wk_sheet in workbook.sheetnames: 
    #     ws = workbook[wk_sheet]
    #     workbook.remove(ws)

    # workbook.save(path)

    return

def save_toExcel_Jyxx_list(text_list,wk_sheet="main",path=PATH+'/result.xlsx'):
    text = []
    for item in text_list:
        item = item.values()
        text.append(item)
    # print(text)
    text = [['发布时间','标题','链接']] +  text
    save_toExcel(text,path=path,wk_sheet=wk_sheet)

def save_toExcel_Jyxx_dict(text_dict,path=PATH+'/result.xlsx'):
    for item in text_dict.items():
        save_toExcel_Jyxx_list(item[1],wk_sheet=item[0],path=path)

def read_fromJson(path=PATH+'/result.json'):
    with open(path, 'r') as f:
        data = json.load(f)
        f.close()
        return data

# ============
# Crawling Related

def get_url(pageIndex, categoryNum = '003002',subCategoryNum='', prt=False):
    # Ease for later modification
    queries = {
        'categoryNum' : str(categoryNum),
        'pageIndex'   :str(pageIndex)
    }
    
    # Generate query sub string 
    query_str = ""
    for item in queries.items():
        query_str += item[0] + '=' + item[1] + '&'
    query_str = query_str[:-1]

    # Render for the final URL
    url = URL_BASE + '/' + URL_SDIR + '/' + URL_FILE + '?' + query_str
    if(prt): print(url)
    return url

def get_url003002(pageIndex):
    if(pageIndex==0): return "http://www.ajztb.com/jyxx/003002/moreinfo.html"
    if(pageIndex==1): return "http://www.ajztb.com/jyxx/003002/moreinfo.html"
    else: return 'http://www.ajztb.com/jyxx/003002/'+str(pageIndex)+'.html'
     
def craw_page(url, header_ = {}, data_ = {}):
    response = requests.get(url, headers=header_, data=data_)
    return response.content.decode('utf-8')

def craw_page_of_date(date, page_from = 1,prt=False,arrayLike=False):

    cur_page = page_from
    debug_from_page = page_from
    debug_to_page   = page_from

    # Retrive to find starting page index by traversing 
    # over the pages 
    found_startPage = False
    while not found_startPage:
        p_src = craw_page(get_url003002(cur_page))
        if(check_containsDate(p_src, date)): found_startPage = True
        elif(check_containsDate(p_src, date) is None): return {}
        else: cur_page+=1
    debug_from_page = cur_page

    # Retrive all page content until next page doesn't contain 
    # any of the specified date 
    found_endPage = False 
    page_source_s = []
    while not found_endPage:
        p_src = craw_page(get_url003002(cur_page))
        if(not check_containsDate(p_src, date)): found_endPage = True
        else: page_source_s.append(p_src)
        cur_page+=1
    
    # DEBUG MSG
    if(prt): 
        debug_to_page = cur_page-2
        print(date.toString(), debug_from_page, debug_to_page)
    if(arrayLike):
        return page_source_s 
    else:
        rtn = ""
        for page_source in page_source_s: rtn += page_source
        return rtn

# =============
# Parsing Related

def extract_pageDates(page_source,prt=False):
    soup = BeautifulSoup(page_source, 'html.parser')
    pageContainer_s = soup.find_all('body')

    rtn_dates = []
    for pageContainer in pageContainer_s:
        dContainer_s = pageContainer.find_all('span',class_='r ewb-notice-date')
        for dContainer in dContainer_s:
            date = dContainer.text      # Could cause encoding error (use contenct.encoding(native_encoding?) encoding if it happends)
            date = date.replace(' ','') # Remove spaces 
            # date = date.replace('/n','')# Remove change line 
            # date = date.replace('/t','')# Remove tab
            rtn_dates.append(date)
            if(prt): print(date)

    return rtn_dates

def check_containsDate(page_source, date):
    date_str = date.toString()
    extracted_dates = extract_pageDates(page_source)

    if(date_str in extracted_dates): return True
    if(len(extracted_dates)==0): return False

    d = extracted_dates[0].split('-')
    l_date = Date(int(d[0]), int(d[1]), int(d[2]))
    compare = date.compareTo(l_date)
    print(date.toString(), l_date.toString(), compare)
    if(compare): return None

    return False

def extract_content_of_date(page_source, date, prt=False):
    soup = BeautifulSoup(page_source, 'html.parser')
    pageContainer_s = soup.find_all('body')

    result_list = []
    for page_container in pageContainer_s:
        content_container = soup.find('div', class_='ewb-list-main')                # Wrapper for the category page
        posts_container   = content_container.find('ul',class_='ewb-notice-items')  # Unordered list  
        posts = posts_container.find_all('li')                                      # List of related items 

        for post in posts:
            title = post.find('a').text
            rel_link = post.find('a')['href']
            link = URL_BASE + rel_link
            date_ = post.find('span',class_='r ewb-notice-date').text
            date_ = date_.replace(' ', '')
            
            if(date.toString() == date_):
                # print("#"*100)
                # print("DEBUG - title: \n\t" + title)     
                # print("DEBUG - link: \n\t" + link)
                # print("DEBUG - date: \n\t" + date_)
                temp_result = {'date':date_,'title':title,'link':link, }

                # if(recursive):                                          # 是否爬取二级网页的内容
                #     url_secondary = temp_result['link']                 # 二级页面链接
                #     res_secondary = craw_page(url_secondary)            # 响应内容
                #     result_secondary = parse_detailPage(res_secondary)  # 解析找到字段
                # else:
                #     result_secondary={}
                # temp_result['detail'] = str(result_secondary).replace('\n','')

                # print(temp_result)
                result_list.append(temp_result) 

    return result_list

# ============
# Combined 

def find_curDate():
    mainPage_url = get_url003002(1)
    page_source  = craw_page(mainPage_url)
    page_dates   = extract_pageDates(page_source)
    cur_date     = (page_dates[0]).split('-')
    
    cur_year    = int(cur_date[0]) 
    cur_month   = int(cur_date[1])
    cur_day     = int(cur_date[2])  

    return Date(cur_year,cur_month,cur_day)

def getPageDate_toExcel(date_from, date_to):
    date_s   = date_to.iter_date_to(date_from)

    result_dict = {}
    for date in date_s:
        source = craw_page_of_date(date)
        if(not len(source)==0):
            result = extract_content_of_date(source,date)
        else:
            result = []
        result_dict[date.toString()] = result

    save_toExcel_Jyxx_dict(result_dict)
    return

def getPageDate_toExcel_crossCheck(date_from, date_to):
    date_s   = date_to.iter_date_to(date_from)

    # Craw all these dates for result 
    result_dict = {}
    for date in date_s:
        source = craw_page_of_date(date)
        if(not len(source)==0):
            result = extract_content_of_date(source,date)
        else:
            result = []
        result_dict[date.toString()] = result

    # Append new date to excel file
    save_toExcel_Jyxx_dict(result_dict)

    # Cross check crawed data with the existing ones 
    found_diff = False
    dict_existing = read_fromJson() # 已有的 POST
    dict_new      = result_dict     # 新发现的 POST
    dict_diff     = {}              # 字典 - 不同
    for date in date_s:
        d_str = date.toString()                  # date that gets checked 
        if(not(date.toString() in list(dict_existing.keys()))):
            found_diff = True
            dict_diff[d_str] = dict_new[d_str]
            dict_existing[d_str] = dict_new[d_str]
            continue 
        list_exist = dict_existing[d_str]
        list_new   = dict_new[d_str]
        dict_existing[d_str] =  list_new         # update new existing (will be added to json file)
        list_diff  = list_new[:-len(list_exist)]
        if(len(list_diff) != 0): 
            found_diff = True
            dict_diff[d_str] = list_diff         # find new posts  
    save_toJson(dict_existing)      # 保存新的JSON文件

    # Notify if required to do so
    if(found_diff):
        clear_Excel(PATH + '/new_data.xlsx')
        save_toExcel_Jyxx_dict(dict_diff, PATH + '/new_data.xlsx')
        
        msg = "\n"
        for i in list(dict_diff.keys()):
            msg += "\t" + i + "\n"
            for j in dict_diff[i]:
                msg += "\t\t-" + j['title'] + '\n'
                msg += "\t\t " + j['link'] + '\n'

        sendEmail163(msg, [PATH + '/new_data.xlsx', PATH + '/result.xlsx'])
        
        print("Sent Mail:\n"+msg)
        # print(dict_diff)
        # print(found_diff)

    return

# ============
# Unit Testing (Mock)

def test_exportMultiHtml():
    saved = ""
    # saved += craw_page(get_url003002(1))   # 2020 12.17 only
    saved += craw_page(get_url003002(3))   # 2020 12.15 to 12.14
    # saved += craw_page(get_url003002(5))   # 2020 12.11 to 12.10
    save_toHtml(saved)
    return saved

def test_formatDatePrinting():
    date = Date(2020, 12, 17)
    date_str = date.toString()
    print(date_str)
    return date_str

def test_extractDates():
    page_source = test_exportMultiHtml()
    result = extract_pageDates(page_source)
    print(result)
    return result

def test_checkPageContainsDate():
    # date = Date(2020, 12, 17)
    # date = Date(2020, 12, 8)
    date = Date(2020, 12, 14)
    page_source = test_exportMultiHtml()
    result = check_containsDate(page_source, date)
    print(result)
    return result

def test_crawDatePages():
    date = Date(2020, 12, 17)
    ps_s = craw_page_of_date(date, prt=True)
    p_ss = ""
    for i in ps_s:
        p_ss += i
    save_toHtml(p_ss)
    return

def test_parseDatePages_n_jsonRead():
    date   = Date(2020,12,17)
    source = craw_page_of_date(date)
    result = extract_content_of_date(source,date)
    save_toJson(result)
    # print(result)
    json_result = read_fromJson()
    print(json_result)
    return result

def test_from_toDate():
    date_from = Date(2020,12,18)
    date_to   = Date(2020,11,19)
    date_s = date_to.iter_date_to(date_from)
    for i in date_s: 
        print(i.toString())
    return 

def test_getData_from_to():
    date_from = Date(2020,12,10)
    date_to   = Date(2020,12,10)
    getPageDate_toExcel(date_from, date_to)
    return 

# ============
# Main / 

def get_specDate(date):
    temp = Date(date.year,date.month,date.day)
    return getPageDate_toExcel_crossCheck(date, temp)

def get_rangeDate(date_from, date_to):
    return getPageDate_toExcel_crossCheck(date_from, date_to)

def get_curDate(prevDaysToUpdate = 1):
    cur_date = find_curDate()
    til_date = Date(cur_date.year, cur_date.month, cur_date.day)
    for x in range(prevDaysToUpdate): til_date.prev()
    return get_rangeDate(cur_date, til_date)
    
def main():

    # get_specDate  (Date(2020,12,17))
    # get_rangeDate (Date(2020,12,18), Date(2020,11,28))
    get_curDate   ()
   
if __name__ == "__main__":
    main()
    